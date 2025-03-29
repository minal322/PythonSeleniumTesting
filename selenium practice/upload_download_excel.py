import time

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl

service_obj = Service("C:/Users/Admin/Downloads/chromedriver-win64/chromedriver-win64/chromedriver.exe")
driver = webdriver.Chrome(service=service_obj)
fruit_name = "Apple"
new_value = "1000"
file_path = "C:/Users/Admin/Downloads/download.xlsx"
dict = { }

driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.maximize_window()
driver.implicitly_wait(2)

def update_excel_data(filepath,fruitname,col_name,updated_price):
    # updating excel
    download_data = openpyxl.load_workbook(filepath)
    sheet = download_data.active
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == col_name:
            dict["col"] = i
    for i in range(1, sheet.max_row + 1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == fruitname:
                dict["row"] = i

    # edit the excel sheet
    sheet.cell(row=dict["row"], column=dict["col"]).value = updated_price
    download_data.save(filepath)

#click on download
driver.find_element(By.ID,"downloadButton").click()

######Once file is downloaded edit file and upload
update_excel_data(file_path,fruit_name,"price",new_value)

#click on upload
element = driver.find_element(By.ID,"fileinput")
element.send_keys(file_path)

#Verify and upload successfully prompt has came up
wait = WebDriverWait(driver,5)
upload_locator = (By.XPATH, "//div[@class='Toastify__toast-body']/div[2]")
wait.until(expected_conditions.visibility_of_element_located(upload_locator))
text = driver.find_element(By.XPATH, "//div[@class='Toastify__toast-body']/div[2]").text
assert "Successfully" in text

#######build dynamic xpath for fruit name and fetch price form web page
#get col id dynamic (means even if new cols add we will get updated id)
price_col_id = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
#xpath from child to parent tag to parent tag to child tag Eg : Apple -> div -> div -> price_vlaue
xpath_price = f"//div[text()={fruit_name}]/parent::div/parent::div/div[@id='cell-{price_col_id}-undefined']"
print(xpath_price)
actual_price = driver.find_element(By.XPATH,f"//div[text()='Apple']/parent::div/parent::div/div[@id='cell-4-undefined']" ).text
print(actual_price)

assert actual_price == new_value

time.sleep(10)