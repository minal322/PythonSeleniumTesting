import openpyxl

book = openpyxl.load_workbook("C:/Users/Admin/Desktop/PNames.xlsx")


sheet = book.active

#to print specific value
cell = sheet.cell(row=1,column=3)
print(cell.value)

#total max row,col present in sheet
print(sheet.max_row)
print(sheet.max_column)

#print vlaue simple way
print(sheet['A3'].value)

#print all values form sheet row and cols
for i in range(1,sheet.max_row +1):
    for j in range(1,sheet.max_column +1 ):
        print(sheet.cell(row=i, column=j).value)

# print values of specific row
for i in range(1,sheet.max_row +1):
    if sheet.cell(row=i,column=1).value == "b":
        for j in range(1,sheet.max_column +1 ):
            print(sheet.cell(row=i, column=j).value)

#Fetch test case 2 data and store in dictionary - [{{"name":'minal','lname':'patil','gender':'female'}, ....}]
dataa = {}
for i in range(1,sheet.max_row +1):
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(2,sheet.max_column +1 ):
           dataa[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value


print(dataa)